import openpyxl as xl
from openpyxl.chart import BarChart,Reference


wb = xl.load_workbook('Inventory.xlsx')
sheet = wb['Sheet1']

cell = sheet.cell(2,2)

for row in range(3, sheet.max_row+1):
    cell = sheet.cell(row,4)
    new_inventory = cell.value + 100
    new_inventory_column = sheet.cell(row,5)
    new_inventory_column.value = new_inventory

    
sheet.cell(2,5).value = "Updated Inventory"
sheet.cell(2,4).value = "Initial Inventory"

Values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=5)

chart = BarChart()
chart.add_data(Values, titles_from_data=True)
chart.title = "Inventory Details"
chart.x_axis.title = "Item"
chart.y_axis.title = "Quantity"

categories = Reference(sheet,
                       min_col=3,
                       max_col=3,
                       min_row=3,
                       max_row=sheet.max_row)
chart.set_categories(categories)

sheet.add_chart(chart,'g3')

wb.save('UpdatedInventory.xlsx')